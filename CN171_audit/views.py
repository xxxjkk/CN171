import csv
import datetime
import os
import re
import time
from io import BytesIO, StringIO

import pandas as pd
from django.db import transaction
from django.db.models import Q
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils.http import urlquote
from openpyxl import Workbook, load_workbook
import json

from CN171_tools.common_api import to_ints

try:
    import ConfigParser as cp
except ImportError as e:
    import configparser as cp

from CN171_background.api import pages, get_object
from CN171_audit.models import staffInfo, hostAccountList, databaseAccountList, systemList, applicationAccountList, \
    privilegedAccountRecord, privilegedAccountWithoutRecord, sensitiveInformation, treasuryAuthenticated

# 读取配置文件
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
config = cp.ConfigParser()
config.read(os.path.join(BASE_DIR, 'config/cn171.conf'), encoding='utf-8')

# 文件路径
audit_file_download = config.get('audit', 'audit_file_download')
audit_file_upload = config.get('audit', 'audit_file_upload')
audit_file_result = config.get('audit', 'audit_file_result')


def infoManagement(request):
    staff_list = staffInfo.objects.all()
    staff_info_list = []
    for staff in staff_list:
        staff_id = staff.staff_id
        staff_name = staff.staff_name
        staff_4AAccount = staff.staff_4AAccount
        staff_account = staff.staff_account
        staff_group = staff.staff_group
        staff_system = staff.staff_system
        staff_post = staff.staff_post
        staff_status = staff.staff_status
        staff_info = {"staff_id": staff_id, "staff_name": staff_name,
                      "staff_4AAccount": staff_4AAccount, "staff_account": staff_account,
                      "staff_group": staff_group, "staff_system": staff_system,
                      "staff_post": staff_post, "staff_status": staff_status}
        staff_info_list.append(staff_info)
    p, page_objects, page_range, current_page, show_first, show_end, end_page, page_len = pages(staff_info_list,
                                                                                                request)
    return render(request, "audit/information.html", locals())


def staffAdd(request):
    if request.method == "POST":
        staff_name = request.POST['staff_name']
        staff_4AAccount = request.POST['staff_4AAccount']
        staff_account = request.POST['staff_account']
        staff_group = request.POST['staff_group']
        staff_system = request.POST['staff_system']
        staff_post = request.POST['staff_post']
        staff_status = request.POST['staff_status']
        staff = staffInfo()
        staff.staff_name = staff_name
        staff.staff_4AAccount = staff_4AAccount
        staff.staff_account = staff_account
        staff.staff_group = staff_group
        staff.staff_system = staff_system
        staff.staff_post = staff_post
        staff.staff_status = staff_status
        staff.save()
        status = 1
        return render(request, "audit/informationAdd.html", locals())
    else:
        display_control = "none"
        return render(request, "audit/informationAdd.html", locals())


def staffMod(request, staff_id):
    staff = get_object(staffInfo, staff_id=staff_id)
    if request.method == "POST":
        staff_name = request.POST['staff_name']
        staff_4AAccount = request.POST['staff_4AAccount']
        staff_account = request.POST['staff_account']
        staff_group = request.POST['staff_group']
        staff_system = request.POST['staff_system']
        staff_post = request.POST['staff_post']
        staff_status = request.POST['staff_status']
        staff.staff_name = staff_name
        staff.staff_4AAccount = staff_4AAccount
        staff.staff_account = staff_account
        staff.staff_group = staff_group
        staff.staff_system = staff_system
        staff.staff_post = staff_post
        staff.staff_status = staff_status
        staff.save()
        status = 1
        return render(request, "audit/informationMod.html", locals())
    else:
        display_control = "none"
        return render(request, "audit/informationMod.html", locals())


def staffDel(request):
    staff_ids = request.POST.getlist('ids', [])
    ret = "True"
    for staff_id in staff_ids:
        try:
            staffInfo.objects.filter(staff_id=staff_id).delete()
        except:
            ret = "False"
    return JsonResponse({'ret': ret})


def importStaffInfo(request):
    ret = {'msg': "上传成功！"}
    file_obj = request.FILES.get('staffInfoFile')
    if file_obj:
        type_excel = file_obj.name.split('.')[1]
        if type_excel in ['xlsx', 'xls']:
            # 开始解析上传的excel表格
            wb = load_workbook(file_obj)
            table = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            nrows = table.max_row
            with transaction.atomic():
                for i in range(2, nrows + 1):
                    name = table.cell(row=i, column=1).value
                    try:
                        staff = staffInfo.objects.get(staff_name=name)
                        staff.staff_name = table.cell(row=i, column=1).value
                        staff.staff_4AAccount = table.cell(row=i, column=2).value
                        staff.staff_account = table.cell(row=i, column=3).value
                        staff.staff_group = table.cell(row=i, column=4).value
                        staff.staff_system = table.cell(row=i, column=5).value
                        staff.staff_post = table.cell(row=i, column=6).value
                        staff.staff_status = table.cell(row=i, column=7).value
                        staff.save()
                    except Exception as e:
                        if str(e) == "staffInfo matching query does not exist.":
                            staff = staffInfo()
                            staff.staff_name = table.cell(row=i, column=1).value
                            staff.staff_4AAccount = table.cell(row=i, column=2).value
                            staff.staff_account = table.cell(row=i, column=3).value
                            staff.staff_group = table.cell(row=i, column=4).value
                            staff.staff_system = table.cell(row=i, column=5).value
                            staff.staff_post = table.cell(row=i, column=6).value
                            staff.staff_status = table.cell(row=i, column=7).value
                            staff.save()
                        else:
                            ret['msg'] = "出现错误：%s" % e
                            break
        else:
            ret['msg'] = "上传文件格式不是xlsx或xls！"
    else:
        ret['msg'] = "请选择上传的文件！"
    v = json.dumps(ret)
    return HttpResponse(v)


def exportStaffInfo(request):
    staff_ids = to_ints(request.GET.get('ids'))
    staff_search = request.GET.get('staff_search')
    wb = Workbook()
    wb.encoding = "utf-8"
    sheet1 = wb.active
    sheet1.title = "部门人员信息"
    row_one = ['姓名', '4A主账号', '从账号', '所在组', '系统', '职责', '状态']
    for i in range(1, len(row_one) + 1):
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    if staff_ids == [-1]:
        staffs = staffInfo.objects.filter(Q(staff_name=staff_search) |
                                          Q(staff_group__icontains=staff_search) |
                                          Q(staff_system__icontains=staff_search))
        for staff in staffs:
            cur_row = sheet1.max_row + 1
            staff_info = [staff.staff_name, staff.staff_4AAccount, staff.staff_account, staff.staff_group,
                          staff.staff_system, staff.staff_post, staff.staff_status]
            for i in range(1, len(staff_info) + 1):
                sheet1.cell(row=cur_row, column=i).value = staff_info[i - 1]
    else:
        for staff_id in staff_ids:
            staff = get_object(staffInfo, staff_id=staff_id)
            cur_row = sheet1.max_row + 1
            staff_info = [staff.staff_name, staff.staff_4AAccount, staff.staff_account, staff.staff_group,
                          staff.staff_system, staff.staff_post, staff.staff_status]
            for i in range(1, len(staff_info) + 1):
                sheet1.cell(row=cur_row, column=i).value = staff_info[i - 1]
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    time = datetime.datetime.now().strftime('%Y-%m-%d')
    file_name = '部门人员信息%s.xlsx' % time
    file_name = urlquote(file_name)
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response


def staffSearch(request):
    staff_search = request.GET.get('staff_search')
    if staff_search == "":
        staff = staffInfo.objects.all()
    else:
        staff = staffInfo.objects.filter(Q(staff_name=staff_search) |
                                         Q(staff_group__icontains=staff_search) |
                                         Q(staff_system__icontains=staff_search))
    p, page_objects, page_range, current_page, show_first, show_end, end_page, page_len = pages(staff, request)
    return render(request, "audit/information.html", locals())


def accountAudit(request):
    systems = systemList.objects.all()
    system_list = []
    for system in systems:
        system_id = system.system_id
        system_name = system.system_name
        system_list.append(system_name)
    system_checked = request.GET.get('system')
    account_num = hostAccountList.objects.filter(host_system=system_checked).count() + \
                  databaseAccountList.objects.filter(database_system=system_checked).count() + \
                  applicationAccountList.objects.filter(application_system=system_checked).count()
    isolated_account_num = hostAccountList.objects.filter(host_system=system_checked, is_isolated="是", is_in_4A="是", is_in_resource="是").count() + \
                           databaseAccountList.objects.filter(database_system=system_checked, is_isolated="是", is_in_4A="是", is_in_resource="是").count() + \
                           applicationAccountList.objects.filter(application_system=system_checked, is_isolated="是", is_in_4A="是", is_in_resource="是").count()
    public_account_num = hostAccountList.objects.filter(host_system=system_checked, is_public="是", is_in_4A="是", is_in_resource="是").count() + \
                         databaseAccountList.objects.filter(database_system=system_checked, is_public="是", is_in_4A="是", is_in_resource="是").count() + \
                         applicationAccountList.objects.filter(application_system=system_checked, is_public="是", is_in_4A="是", is_in_resource="是").count()
    isolated_public_num = isolated_account_num + public_account_num
    inconsistent_account_num = hostAccountList.objects.filter(host_system=system_checked, is_in_4A="是", is_in_resource="否").count() + \
                               hostAccountList.objects.filter(host_system=system_checked, is_in_4A="否", is_in_resource="是").count() + \
                               databaseAccountList.objects.filter(database_system=system_checked, is_in_4A="是", is_in_resource="否").count() + \
                               databaseAccountList.objects.filter(database_system=system_checked, is_in_4A="否", is_in_resource="是").count() + \
                               applicationAccountList.objects.filter(application_system=system_checked, is_in_4A="是", is_in_resource="否").count() + \
                               applicationAccountList.objects.filter(application_system=system_checked, is_in_4A="否", is_in_resource="是").count()
    change_host_group_num = hostAccountList.objects.filter(Q(host_system=system_checked) & Q(how_to_change__icontains="变更主组")).count()
    change_database_role_num = databaseAccountList.objects.filter(Q(database_system=system_checked) & Q(how_to_change__icontains="变更角色权限")).count()
    change_authority_num = change_host_group_num + change_database_role_num
    change_account_type_num = hostAccountList.objects.filter(Q(host_system=system_checked) & Q(how_to_change__icontains="变更账号属性")).count() + \
                              databaseAccountList.objects.filter(Q(database_system=system_checked) & Q(how_to_change__icontains="变更账号属性")).count() + \
                              applicationAccountList.objects.filter(Q(application_system=system_checked) & Q(how_to_change__icontains="变更账号属性")).count()
    miss_delete_account_num = hostAccountList.objects.filter(host_system=system_checked, is_keep="否").count() + \
                              databaseAccountList.objects.filter(database_system=system_checked, is_keep="否").count() + \
                              applicationAccountList.objects.filter(application_system=system_checked, is_keep="否").count()
    account_to_cancel_num = hostAccountList.objects.filter(host_system=system_checked, is_to_cancel="是").count() + \
                            databaseAccountList.objects.filter(database_system=system_checked, is_to_cancel="是").count() + \
                            applicationAccountList.objects.filter(application_system=system_checked, is_to_cancel="是").count()
    account_to_delete_num = miss_delete_account_num + account_to_cancel_num
    new_without_application_num = hostAccountList.objects.filter(host_system=system_checked, is_keep="是", is_applied="否", is_to_cancel="否").count() + \
                                  databaseAccountList.objects.filter(database_system=system_checked, is_keep="是", is_applied="否", is_to_cancel="否").count() + \
                                  applicationAccountList.objects.filter(application_system=system_checked, is_keep="是", is_applied="否", is_to_cancel="否").count()
    try:
        log_num = open(audit_file_result + "log_num.txt")
        line = log_num.readline()
        log_num_sum = int(line.split(",")[0])
        log_num_without_application = int(line.split(",")[1])
        log_num_without_authentication = int(line.split(",")[2])
    except:
        log_num_sum = 0
        log_num_without_application = 0
    pass_num1 = hostAccountList.objects.filter(host_system=system_checked, is_isolated="否", is_public="否", is_in_4A="是", is_in_resource="是").count() + \
                databaseAccountList.objects.filter(database_system=system_checked, is_isolated="否", is_public="否", is_in_4A="是", is_in_resource="是").count() + \
                applicationAccountList.objects.filter(application_system=system_checked, is_isolated="否", is_public="否", is_in_4A="是", is_in_resource="是").count() + \
                hostAccountList.objects.filter(host_system=system_checked, is_isolated="否", is_public="否", is_in_4A="否", is_in_resource="否").count() + \
                databaseAccountList.objects.filter(database_system=system_checked, is_isolated="否", is_public="否", is_in_4A="否", is_in_resource="否").count() + \
                applicationAccountList.objects.filter(application_system=system_checked, is_isolated="否", is_public="否", is_in_4A="否", is_in_resource="否").count()
    try:
        pass_percent1 = round(pass_num1 / account_num, 2)
    except:
        pass_percent1 = 1
    pass_percent1 = int(pass_percent1 * 100)
    pass_num2 = hostAccountList.objects.filter(host_system=system_checked, is_to_change="否").count() + \
                databaseAccountList.objects.filter(database_system=system_checked, is_to_change="否").count() + \
                applicationAccountList.objects.filter(application_system=system_checked, is_to_change="否").count()
    try:
        pass_percent2 = round(pass_num2 / account_num, 2)
    except:
        pass_percent2 = 1
    pass_percent2 = int(pass_percent2 * 100)
    pass_num3 = hostAccountList.objects.filter(host_system=system_checked, is_to_cancel="否", is_keep="是", is_applied="是").count() + \
                databaseAccountList.objects.filter(database_system=system_checked, is_to_cancel="否", is_keep="是", is_applied="是").count() + \
                applicationAccountList.objects.filter(application_system=system_checked, is_to_cancel="否", is_keep="是", is_applied="是").count()
    try:
        pass_percent3 = round(pass_num3 / account_num, 2)
    except:
        pass_percent3 = 1
    pass_percent3 = int(pass_percent3 * 100)
    pass_num4 = log_num_sum - log_num_without_application - log_num_without_authentication
    try:
        pass_percent4 = round(pass_num4 / log_num_sum, 2)
    except:
        pass_percent4 = 1
    pass_percent4 = int(pass_percent4 * 100)
    return render(request, "audit/accountAudit.html", locals())


def systemAdd(request):
    if request.method == "POST":
        try:
            system_name = request.POST['system_name']
            if system_name == "":
                status = 2
                return render(request, "audit/systemAdd.html", locals())
            else:
                system = systemList()
                system.system_name = system_name
                system.save()
                status = 1
                return render(request, "audit/systemAdd.html", locals())
        except:
            status = 2
            return render(request, "audit/systemAdd.html", locals())
    else:
        display_control = "none"
        return render(request, "audit/systemAdd.html", locals())


def systemDel(request):
    system = request.POST.get('system')
    ret = "True"
    try:
        systemList.objects.filter(system_name=system).delete()
    except:
        ret = "False"
    return JsonResponse({'ret': ret})


def download(request, file_id):
    files = {'1': "开户+变更+销户记录模板.xlsx",
             '2': "4A主账号与从账号对应清单模板.xlsx",
             '3': "资源侧主机从账号信息模板.txt",
             '4': "特权+程序账号备案记录模板.xlsx",
             '5': "操作日志模板（北京）.csv",
             '6': "操作日志模板（深圳）.csv",
             '7': "资源侧数据库从账号信息模板.txt",
             '8': "金库场景.xlsx",
             '9': "金库审批报表.xlsx"}
    file_name = files[file_id]
    file = open(audit_file_download + file_name, "rb")
    response = FileResponse(file)
    response['Content-Type'] = "application/octet-stream"
    response['Content-Disposition'] = "attachment;filename=" + file_name
    return response


def upload(request, file_id):
    types = {'1': ["开户+变更+销户记录.xlsx", "xlsx", ["开户", "变更", "销户"]],
             '2': ["4A主账号与从账号对应清单.xlsx", "xlsx",
                   ["主机账号列表（北京4A）", "数据库账号列表（北京4A）", "前台账号列表（北京4A）",
                    "主机+数据库账号列表（深圳4A）", "前台账号列表（深圳4A）", "说明"]],
             '3': ["资源侧主机从账号信息.txt", "txt"],
             '4': ["特权+程序账号备案记录.xlsx", "xlsx", ["特权+程序账号备案记录"]],
             '5': ["操作日志.csv", "csv", ["操作日志（北京4A）", "操作日志（深圳4A）"]],
             '6': ["资源侧数据库从账号信息.txt", "txt"],
             '7': ["金库场景.xlsx", "xlsx", ["金库场景"]],
             '8': ["金库审批报表.xlsx", "xlsx", ["金库审批报表"]]}
    ret = {'msg': "上传成功！"}
    system = request.POST.get('system')
    file_num = request.POST.get('fileNum')
    if file_num != 0:
        type = request.FILES.get('accountFile[0]').name.split('.')[1]
        if type == types[file_id][1]:
            if type == "xlsx":
                if file_id == "4":
                    for i in range(0, int(file_num)):
                        file_obj = request.FILES.get('accountFile[' + str(i) + ']')
                        wb = load_workbook(file_obj)
                        table_names = wb.get_sheet_names()
                        for table_name in table_names:
                            table = wb.get_sheet_by_name(table_name)
                            nrows = table.max_row
                            ncolumns = table.max_column
                            columns = {'系统名称': "-1",
                                       '资源ip': "-1",
                                       '资源名称': "-1",
                                       '资源类型': "-1",
                                       '操作账号': "-1",
                                       '程序账号': "-1",
                                       '操作时间': "-1"}
                            for i in range(1, ncolumns + 1):
                                if table.cell(row=1, column=i).value in ['业务系统']:
                                    columns['系统名称'] = i
                                elif table.cell(row=1, column=i).value in ['资源ip']:
                                    columns['资源ip'] = i
                                elif table.cell(row=1, column=i).value in ['资源名称']:
                                    columns['资源名称'] = i
                                elif table.cell(row=1, column=i).value in ['操作类型']:
                                    columns['资源类型'] = i
                                elif table.cell(row=1, column=i).value in ['操作账号', '操作账号(主账号/从账号）']:
                                    columns['操作账号'] = i
                                elif table.cell(row=1, column=i).value in ['申请程序账号']:
                                    columns['程序账号'] = i
                                elif table.cell(row=1, column=i).value in ['操作时间']:
                                    columns['操作时间'] = i
                            for i in range(2, nrows + 1):
                                print(table.cell(row=i, column=int(columns['系统名称'])).value)
                                if re.search(system, table.cell(row=i, column=int(columns['系统名称'])).value, re.IGNORECASE):
                                    system_name = system
                                    resource_ip = table.cell(row=i, column=int(columns['资源ip'])).value
                                    resource_name = table.cell(row=i, column=int(columns['资源名称'])).value
                                    resource_type = table.cell(row=i, column=int(columns['资源类型'])).value
                                    operating_account = table.cell(row=i, column=int(columns['操作账号'])).value
                                    operating_time = str(table.cell(row=i, column=int(columns['操作时间'])).value)
                                    if system in ["内容计费", "物漫"]:
                                        personal_account = operating_account
                                        privilaged_account = table.cell(row=i, column=int(columns['程序账号'])).value
                                    else:
                                        accounts = operating_account.replace("\n", "").split("/")
                                        personal_account = accounts[0]
                                        privilaged_account = accounts[1]
                                    try:
                                        times = operating_time.replace("\n", "").split("-")
                                        date = times[0].split(".")
                                        if len(date) == 3:
                                            year = date[0]
                                        else:
                                            year = str(datetime.datetime.now().year)
                                        month = date[len(date) - 2]
                                        day = date[len(date) - 1]
                                        times[0] = year + "." + month + "." + day
                                        if len(times) == 1:
                                            start_time = int(time.mktime(time.strptime(times[0], '%Y.%m.%d')))
                                            end_time = start_time + 86399
                                        else:
                                            if len(times[0].split(" ")) == 1:
                                                start_time = int(time.mktime(time.strptime(times[0], '%Y.%m.%d')))
                                            else:
                                                start_time = int(time.mktime(time.strptime(times[0], '%Y.%m.%d %H:%M:%S')))
                                            if len(times[1].split(" ")) == 1:
                                                if len(times[1].split(".")) == 2:
                                                    times[1] = year + "." + times[1]
                                                elif len(times[1].split(".")) == 1:
                                                    times[1] = year + "." + month + "." + times[1]
                                                end_time = int(time.mktime(time.strptime(times[1], '%Y.%m.%d'))) + 86399
                                            else:
                                                end_time = int(time.mktime(time.strptime(times[1], '%Y.%m.%d %H:%M:%S')))
                                        privilaged_account_record = privilegedAccountRecord()
                                        privilaged_account_record.system_name = system_name
                                        privilaged_account_record.resource_ip = resource_ip
                                        privilaged_account_record.resource_name = resource_name
                                        privilaged_account_record.resource_type = resource_type
                                        privilaged_account_record.personal_account = personal_account
                                        privilaged_account_record.privileged_account = privilaged_account
                                        privilaged_account_record.start_time = start_time
                                        privilaged_account_record.end_time = end_time
                                        privilaged_account_record.save()
                                    except:
                                        import traceback
                                        traceback.print_exc()
                                        ret['msg'] = file_obj.name + "的第" + str(i) + "行字段格式不正确！"
                elif file_id == "7":
                    file_obj = request.FILES.get('accountFile[0]')
                    wb = load_workbook(file_obj)
                    table_name = wb.get_sheet_names()
                    table = wb.get_sheet_by_name(table_name[0])
                    nrows = table.max_row
                    ncolumns = table.max_column
                    columns = {'敏感场景': "-1",
                               '资源类型': "-1",
                               '系统名称': "-1"}
                    for i in range(1, ncolumns + 1):
                        if table.cell(row=1, column=i).value in ['敏感表/文件/页面']:
                            columns['敏感场景'] = i
                        elif table.cell(row=1, column=i).value in ['资源类型']:
                            columns['资源类型'] = i
                        elif table.cell(row=1, column=i).value in ['系统']:
                            columns['系统名称'] = i
                    for i in range(2, nrows + 1):
                        print(table.cell(row=i, column=int(columns['系统名称'])).value)
                        if re.search(system, table.cell(row=i, column=int(columns['系统名称'])).value, re.IGNORECASE):
                            system_name = system
                            resource_type = table.cell(row=i, column=int(columns['资源类型'])).value
                            sensitive_item = table.cell(row=i, column=int(columns['敏感场景'])).value
                            sensitive_scene = sensitiveInformation()
                            sensitive_scene.sensitive_item = sensitive_item
                            sensitive_scene.resource_type = resource_type
                            sensitive_scene.system = system_name
                            sensitive_scene.save()
                elif system in ["CMIOT", "BBOSS", "内容计费"] and file_id == "8":
                    file_obj = request.FILES.get('accountFile[0]')
                    wb = load_workbook(file_obj)
                    table_names = wb.get_sheet_names()
                    for table_name in table_names:
                        table = wb.get_sheet_by_name(table_name)
                        nrows = table.max_row
                        ncolumns = table.max_column
                        columns = {'系统名称': "-1",
                                   '操作账号': "-1",
                                   '操作内容': "-1",
                                   '申请开始时间': "-1",
                                   '申请结束时间': "-1"}
                        for i in range(1, ncolumns + 1):
                            if table.cell(row=1, column=i).value in ['申请人组织机构']:
                                columns['系统名称'] = i
                            elif table.cell(row=1, column=i).value in ['申请人从帐号']:
                                columns['操作账号'] = i
                            elif table.cell(row=1, column=i).value in ['操作内容']:
                                columns['操作内容'] = i
                            elif table.cell(row=1, column=i).value in ['金库申请时间']:
                                columns['申请开始时间'] = i
                            elif table.cell(row=1, column=i).value in ['金库申请失效时间']:
                                columns['申请结束时间'] = i
                        for i in range(2, nrows + 1):

                            system_name = table.cell(row=i, column=int(columns['系统名称'])).value
                            pattern1 = re.compile(r'实例名[为]*[a-zA-Z0-9\_]+的')
                            print(table.cell(row=i, column=int(columns['操作内容'])).value)
                            resource_name = re.findall(pattern1, table.cell(row=i, column=int(columns['操作内容'])).value)[0]
                            if re.search('为', resource_name):
                                resource_name = resource_name[4:(len(resource_name) - 1)]
                            else:
                                resource_name = resource_name[3:(len(resource_name) - 1)]
                            operating_account = table.cell(row=i, column=int(columns['操作账号'])).value
                            pattern2 = re.compile(r'如下内容表[a-zA-Z0-9\_表]+')
                            sensitive_table = re.findall(pattern2, table.cell(row=i, column=int(columns['操作内容'])).value)[0]
                            sensitive_table = sensitive_table[4:]
                            applied_start_time = table.cell(row=i, column=int(columns['申请开始时间'])).value
                            applied_end_time = table.cell(row=i, column=int(columns['申请结束时间'])).value
                            try:
                                start_time = int(time.mktime(time.strptime(applied_start_time, '%Y-%m-%d %H:%M:%S')))
                                if applied_end_time:
                                    end_time = int(time.mktime(time.strptime(applied_end_time, '%Y-%m-%d %H:%M:%S')))
                                else:
                                    applied_end_time = applied_start_time[0:11] + "23:59:59"
                                    end_time = int(time.mktime(time.strptime(applied_end_time, '%Y-%m-%d %H:%M:%S')))
                                treasury_authenticated = treasuryAuthenticated()
                                treasury_authenticated.resource_name = resource_name
                                treasury_authenticated.operating_account = operating_account
                                treasury_authenticated.sensitive_resource = sensitive_table
                                treasury_authenticated.start_time = start_time
                                treasury_authenticated.end_time = end_time
                                treasury_authenticated.system = system_name
                                treasury_authenticated.save()
                            except:
                                import traceback
                                traceback.print_exc()
                                ret['msg'] = file_obj.name + "的第" + str(i) + "行字段格式不正确！"
                else:
                    file_obj = request.FILES.get('accountFile[0]')
                    wb = load_workbook(file_obj)
                    wb.save(audit_file_upload + types[file_id][0])
            elif type == "csv":
                for i in range(0, int(file_num)):
                    file_obj = request.FILES.get('accountFile[' + str(i) + ']')
                    csv_file = csv.reader(StringIO(file_obj.read().decode('ansi')))
                    csv_header = next(csv_file)
                    if i == 0:
                        result_file = open(audit_file_upload + types[file_id][0], 'w', newline='')
                    else:
                        time.sleep(5)
                        result_file = open(audit_file_upload + types[file_id][0], 'a', newline='')
                    csv_result = csv.writer(result_file)
                    if i == 0:
                        csv_result.writerow(csv_header)
                    for row in csv_file:
                        try:
                            print(row)
                            csv_result.writerow(row)
                        except:
                            continue
                    result_file.close()
                result_file.close()
                frame = pd.read_csv(audit_file_upload + types[file_id][0], engine='python', encoding='ansi')
                data = frame.drop_duplicates()
                data.to_csv(audit_file_upload + types[file_id][0], encoding='ansi', index=False)
            else:
                file_obj = request.FILES.get('accountFile[0]')
                pattern = re.compile(r'[(](.*?)[)]')
                if file_id == "3":
                    hostAccountList.objects.filter(is_in_4A="是", host_system=system).update(account_type="")
                    hostAccountList.objects.filter(is_in_4A="是", host_system=system).update(account_4AAccount="")
                    hostAccountList.objects.filter(is_keep="否", host_system=system).update(is_keep="是")
                    hostAccountList.objects.filter(is_in_4A="是", host_system=system).update(is_in_4A="否")
                    hostAccountList.objects.filter(is_in_resource="是", host_system=system).update(is_in_resource="否")
                    hostAccountList.objects.filter(is_isolated="否", host_system=system).update(is_isolated="是")
                    hostAccountList.objects.filter(is_public="是", host_system=system).update(is_public="否")
                    # hostAccountList.objects.filter(is_applied="否", host_system=system).update(is_applied="是")
                    hostAccountList.objects.filter(is_to_cancel="是", host_system=system).update(is_to_cancel="否")
                    hostAccountList.objects.filter(is_to_change="是", host_system=system).update(is_to_change="否")
                    hostAccountList.objects.filter(is_to_change="是", host_system=system).update(how_to_change="")
                    for line in file_obj.readlines():
                        items = line.decode("utf-8").strip().split()
                        if len(items) == 1 and len(items[0].split(".")) == 4:
                            host_ip = items[0]
                        elif len(items) == 1 and items[0] != "\"\c" and items[0] != "\"":
                            host_name = items[0]
                        elif len(items) > 1:
                            account = re.findall(pattern, items[0])
                            group = re.findall(pattern, items[1])
                            account_detail_host = hostAccountList.objects.filter(account_name=account[0],
                                                                                 host_name=host_name)
                            if account_detail_host:
                                account_detail_host = hostAccountList.objects.get(account_name=account[0],
                                                                                  host_name=host_name)
                                account_detail_host.is_new = "否"
                                account_detail_host.is_applied = "是"
                            else:
                                account_detail_host = hostAccountList()
                                account_detail_host.is_new = "是"
                                account_detail_host.is_keep = "是"
                                account_detail_host.is_in_4A = "否"
                                account_detail_host.is_isolated = "是"
                                account_detail_host.is_public = "否"
                                account_detail_host.is_applied = "否"
                                account_detail_host.is_to_cancel = "否"
                                account_detail_host.host_ip = host_ip
                                account_detail_host.host_name = host_name
                                account_detail_host.account_name = account[0]
                            if staffInfo.objects.filter(Q(staff_account=account[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account[0]) & Q(staff_system__icontains=system[0])):
                                staff_post = staffInfo.objects.get(Q(staff_account=account[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account[0]) & Q(staff_system__icontains=system[0])).staff_post
                                if staff_post == "运维" and system in ["PBOSS", "物漫"]:
                                    if group[0] in ["pbmaintain", "trssadm"]:
                                        account_detail_host.is_to_change = "否"
                                        account_detail_host.how_to_change = ""
                                    else:
                                        account_detail_host.is_to_change = "是"
                                        account_detail_host.how_to_change = "变更主组为pbmaintain/trssadm"
                                else:
                                    if group[0] in ["users", "user"]:
                                        account_detail_host.is_to_change = "否"
                                        account_detail_host.how_to_change = ""
                                    else:
                                        account_detail_host.is_to_change = "是"
                                        account_detail_host.how_to_change = "变更主组为users/user组"
                            else:
                                account_detail_host.is_to_change = "否"
                                account_detail_host.how_to_change = ""
                            account_detail_host.account_group = group[0]
                            account_detail_host.is_in_resource = "是"
                            account_detail_host.host_system = system
                            account_detail_host.save()
                    # hostAccountList.objects.filter(is_keep="否").delete()
                elif file_id == "6":
                    account_name = ""
                    databaseAccountList.objects.filter(is_in_4A="是", database_system=system).update(account_type="")
                    databaseAccountList.objects.filter(is_in_4A="是", database_system=system).update(account_4AAccount="")
                    databaseAccountList.objects.filter(is_keep="否", database_system=system).update(is_keep="是")
                    databaseAccountList.objects.filter(is_in_4A="是", database_system=system).update(is_in_4A="否")
                    databaseAccountList.objects.filter(is_in_resource="是", database_system=system).update(is_in_resource="否")
                    databaseAccountList.objects.filter(is_isolated="否", database_system=system).update(is_isolated="是")
                    databaseAccountList.objects.filter(is_public="是", database_system=system).update(is_public="否")
                    # databaseAccountList.objects.filter(is_applied="否", database_system=system).update(is_applied="是")
                    databaseAccountList.objects.filter(is_to_cancel="是", database_system=system).update(is_to_cancel="否")
                    databaseAccountList.objects.filter(is_to_change="是", database_system=system).update(is_to_change="否")
                    databaseAccountList.objects.filter(is_to_change="是", database_system=system).update(how_to_change="")
                    for line in file_obj.readlines():
                        items = line.decode("utf-8").strip().split()
                        if len(items) == 1 and len(items[0].split(".")) == 4 and not re.search("_", items[0]):
                            database_ip = items[0]
                        elif len(items) == 1 and items[0] != "\"\c" and items[0] != "\"":
                            database_name = items[0]
                        elif len(items) > 1:
                            if account_name == items[0]:
                                account_detail_database.account_role = account_detail_database.account_role + "+" + \
                                                                       items[1]
                            elif account_name == "":
                                account_detail_database = databaseAccountList.objects.filter(account_name=items[0],
                                                                                             database_name=database_name)
                                if account_detail_database:
                                    account_detail_database = databaseAccountList.objects.get(account_name=items[0],
                                                                                              database_name=database_name)
                                    account_detail_database.is_new = "否"
                                    account_detail_database.is_applied = "是"
                                else:
                                    account_detail_database = databaseAccountList()
                                    account_detail_database.is_new = "是"
                                    account_detail_database.is_keep = "是"
                                    account_detail_database.is_in_4A = "否"
                                    account_detail_database.is_isolated = "是"
                                    account_detail_database.is_public = "否"
                                    account_detail_database.is_applied = "否"
                                    account_detail_database.is_to_cancel = "否"
                                    account_detail_database.database_ip = database_ip
                                    account_detail_database.database_name = database_name
                                    account_detail_database.account_name = items[0]
                                account_detail_database.account_role = items[1]
                                account_detail_database.is_in_resource = "是"
                                account_detail_database.database_system = system
                                account_name = items[0]
                            else:
                                if staffInfo.objects.filter(Q(staff_account=items[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=items[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + items[0]) & Q(staff_system__icontains=system[0])):
                                    staff_post = staffInfo.objects.get(Q(staff_account=items[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=items[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + items[0]) & Q(staff_system__icontains=system[0])).staff_post
                                    if staff_post == "运维":
                                        if account_detail_database.account_role == "CMSZ_ROLEA+CMSZ_ROLEOL":
                                            account_detail_database.is_to_change = "否"
                                            account_detail_database.how_to_change = ""
                                        else:
                                            account_detail_database.is_to_change = "是"
                                            account_detail_database.how_to_change = "变更角色权限为CMSZ_ROLEA+CMSZ_ROLEOL"
                                    else:
                                        if account_detail_database.account_role == "CMSZ_ROLEA":
                                            account_detail_database.is_to_change = "否"
                                        else:
                                            account_detail_database.is_to_change = "是"
                                            account_detail_database.how_to_change = "变更角色权限为CMSZ_ROLEA"
                                else:
                                    account_detail_database.is_to_change = "否"
                                    account_detail_database.how_to_change = ""
                                account_detail_database.save()
                                account_detail_database = databaseAccountList.objects.filter(account_name=items[0],
                                                                                             database_name=database_name)
                                if account_detail_database:
                                    account_detail_database = databaseAccountList.objects.get(account_name=items[0],
                                                                                              database_name=database_name)
                                    account_detail_database.is_new = "否"
                                    account_detail_database.is_applied = "是"
                                else:
                                    account_detail_database = databaseAccountList()
                                    account_detail_database.is_new = "是"
                                    account_detail_database.is_keep = "是"
                                    account_detail_database.is_in_4A = "否"
                                    account_detail_database.is_isolated = "是"
                                    account_detail_database.is_public = "否"
                                    account_detail_database.is_applied = "否"
                                    account_detail_database.is_to_cancel = "否"
                                    account_detail_database.database_ip = database_ip
                                    account_detail_database.database_name = database_name
                                    account_detail_database.account_name = items[0]
                                account_detail_database.account_role = items[1]
                                account_detail_database.is_in_resource = "是"
                                account_detail_database.database_system = system
                                account_name = items[0]
                    if staffInfo.objects.filter(Q(staff_account=items[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=items[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + items[0]) & Q(staff_system__icontains=system[0])):
                        staff_post = staffInfo.objects.get(Q(staff_account=items[0]) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=items[0] + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + items[0]) & Q(staff_system__icontains=system[0])).staff_post
                        if staff_post == "运维":
                            if account_detail_database.account_role == "CMSZ_ROLEA+CMSZ_ROLEOL":
                                account_detail_database.is_to_change = "否"
                                account_detail_database.how_to_change = ""
                            else:
                                account_detail_database.is_to_change = "是"
                                account_detail_database.how_to_change = "变更角色权限为CMSZ_ROLEA+CMSZ_ROLEOL"
                        else:
                            if account_detail_database.account_role == "CMSZ_ROLEA":
                                account_detail_database.is_to_change = "否"
                            else:
                                account_detail_database.is_to_change = "是"
                                account_detail_database.how_to_change = "变更角色权限为CMSZ_ROLEA"
                    else:
                        account_detail_database.is_to_change = "否"
                        account_detail_database.how_to_change = ""
                    account_detail_database.save()
                    # databaseAccountList.objects.filter(is_keep="否").delete()
        else:
            ret['msg'] = "上传文件格式不正确！"
    else:
        import traceback
        traceback.print_exc()
        ret['msg'] = "请选择上传的文件！"
    v = json.dumps(ret)
    return HttpResponse(v)


def accountAuditing(request):
    ret = {'msg': "审计完成！"}
    system = request.POST.get("system")
    if system == "CMIOT":
        system = ["CMIOT", "CTBOSS系统", "CTBOSS物联网"]
    elif system == "内容计费":
        system = ["内容计费", "物联网-内容计费"]
    elif system == "物漫":
        system = ["物漫", "物漫系统"]
    else:
        system = [system]
    if system[0] == "PBOSS":
        matchList = ['ssh', 'sudo', 'root@', 'I2000@', 'pboss@', 'dds@', 'oracle@', 'I@', 'P@']
    else:
        matchList = ['sudo', 'root', 'sys', 'system', 'dba']
    try:
        wb = load_workbook(audit_file_upload + "4A主账号与从账号对应清单.xlsx")
        table_names = wb.get_sheet_names()
        for table_name in table_names:
            table = wb.get_sheet_by_name(table_name)
            nrows = table.max_row
            ncolumns = table.max_column
            columns = {'4A主账号': "-1",
                       '姓名': "-1",
                       '从账号': "-1",
                       '从账号类型': "-1",
                       '资源名称': "1",
                       '资源IP': "1",
                       '资源类型': "1",
                       '资源所属业务系统': "-1",
                       '从账号状态': "-1"}
            for i in range(1, ncolumns + 1):
                if table.cell(row=1, column=i).value in ['授权人主账号', '归属主帐号']:
                    columns['4A主账号'] = i
                elif table.cell(row=1, column=i).value in ['授权主账号真实姓名', '姓名', '主帐号姓名', '负责人']:
                    columns['姓名'] = i
                elif table.cell(row=1, column=i).value in ['从账号名', '从帐号名', '应用从帐号名称', '从帐号名称', '从账号名称']:
                    columns['从账号'] = i
                elif table.cell(row=1, column=i).value in ['从账号类型', '从帐号类型', '帐号类型']:
                    columns['从账号类型'] = i
                elif table.cell(row=1, column=i).value in ['资源名称', '主机名称', '数据库名称']:
                    columns['资源名称'] = i
                elif table.cell(row=1, column=i).value in ['资源IP', 'IP地址']:
                    columns['资源IP'] = i
                elif table.cell(row=1, column=i).value in ['资源类型']:
                    columns['资源类型'] = i
                elif table.cell(row=1, column=i).value in ['资源所属业务系统', '业务系统', '应用系统', '所属业务系统', '业务系统名称']:
                    columns['资源所属业务系统'] = i
                elif table.cell(row=1, column=i).value in ['从帐号状态', '应用帐号状态']:
                    columns['从账号状态'] = i
            if int(columns['资源名称']) == int(columns['资源IP']):
                applicationAccountList.objects.filter(is_in_4A="是", application_system=system[0]).update(account_type="")
                applicationAccountList.objects.filter(is_in_4A="是", application_system=system[0]).update(account_4AAccount="")
                applicationAccountList.objects.filter(is_keep="否", application_system=system[0]).update(is_keep="是")
                applicationAccountList.objects.filter(is_in_4A="是", application_system=system[0]).update(is_in_4A="否")
                # applicationAccountList.objects.filter(is_in_resource="是", application_system=system[0]).update(is_in_resource="否")
                applicationAccountList.objects.filter(is_isolated="否", application_system=system[0]).update(is_isolated="是")
                applicationAccountList.objects.filter(is_public="是", application_system=system[0]).update(is_public="否")
                applicationAccountList.objects.filter(is_to_cancel="是", application_system=system[0]).update(is_to_cancel="否")
                applicationAccountList.objects.filter(is_to_change="是", application_system=system[0]).update(is_to_change="否")
                applicationAccountList.objects.filter(is_to_change="是", application_system=system[0]).update(how_to_change="")
            for i in range(2, nrows + 1):
                account_4AAccount = table.cell(row=i, column=int(columns['4A主账号'])).value
                staff_name = table.cell(row=i, column=int(columns['姓名'])).value
                account_name = table.cell(row=i, column=int(columns['从账号'])).value
                account_type = table.cell(row=i, column=int(columns['从账号类型'])).value
                resource_name = table.cell(row=i, column=int(columns['资源名称'])).value
                resource_ip = table.cell(row=i, column=int(columns['资源IP'])).value
                resource_system = table.cell(row=i, column=int(columns['资源所属业务系统'])).value
                if table.cell(row=i, column=int(columns['资源所属业务系统'])).value in system and \
                   (int(columns['从账号状态']) == -1 or table.cell(row=i, column=int(columns['从账号状态'])).value == "正常"):
                    if table.cell(row=i, column=int(columns['资源类型'])).value == "unix" or table.cell(row=1, column=int(
                            columns['资源名称'])).value == "主机名称":
                        account_detail_host = hostAccountList.objects.filter(account_name=account_name,
                                                                             host_name=resource_name)
                        if account_detail_host:
                            account_detail_host = hostAccountList.objects.get(account_name=account_name,
                                                                              host_name=resource_name)
                            if account_detail_host.is_in_resource == "否":
                                account_detail_host.is_new = "否"
                                account_detail_host.is_applied = "是"
                        else:
                            account_detail_host = hostAccountList.objects.filter(account_4AAccount=account_4AAccount,
                                                                                 account_name="",
                                                                                 host_name=resource_name)
                            if account_detail_host:
                                account_detail_host = hostAccountList.objects.get(account_4AAccount=account_4AAccount,
                                                                                  account_name="",
                                                                                  host_name=resource_name)
                            else:
                                account_detail_host = hostAccountList()
                                account_detail_host.is_new = "是"
                                account_detail_host.is_applied = "否"
                            account_detail_host.account_name = account_name
                            account_detail_host.host_name = resource_name
                            account_detail_host.host_ip = resource_ip
                            account_detail_host.is_in_resource = "否"
                        if account_detail_host.account_4AAccount in [None, ""] or account_detail_host.account_4AAccount == account_4AAccount:
                            account_detail_host.account_4AAccount = account_4AAccount
                            account_detail_host.is_public = "否"
                        else:
                            account_detail_host.account_4AAccount = account_detail_host.account_4AAccount + "," + account_4AAccount
                            account_detail_host.is_public = "是"
                        if account_type in ["用户具名账号", "普通帐号"] and \
                           staffInfo.objects.filter(Q(staff_name=staff_name) &
                                                    Q(staff_system__icontains=system[0]) &
                                                    Q(staff_status="在职")):
                            account_detail_host.is_to_cancel = "否"
                        elif account_type not in ["用户具名账号", "普通帐号"]:
                            account_detail_host.is_to_cancel = "否"
                            if account_name + "@" not in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                matchList.append(account_name + "@")
                        else:
                            account_detail_host.is_to_cancel = "是"
                        if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                            account_detail_host.staff_name = staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_name
                            if account_type not in ["用户具名账号", "普通帐号"]:
                                if account_name + "@" in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                    matchList.remove(account_name + "@")
                                account_detail_host.is_to_change = "是"
                                if account_detail_host.how_to_change == "":
                                    account_detail_host.how_to_change = "变更账号属性为用户具名账号/普通帐号"
                                elif not re.search("账号属性", account_detail_host.how_to_change):
                                    account_detail_host.how_to_change = account_detail_host.how_to_change + "+变更账号属性为用户具名账号/普通帐号"
                        else:
                            account_detail_host.staff_name = staff_name
                        account_detail_host.account_type = account_type
                        account_detail_host.host_system = system[0]
                        account_detail_host.is_keep = "是"
                        account_detail_host.is_in_4A = "是"
                        if account_detail_host.account_4AAccount in [None, ""] and account_detail_host.is_in_4A == "是":
                            account_detail_host.is_isolated = "是"
                        else:
                            account_detail_host.is_isolated = "否"
                        account_detail_host.save()
                    elif table.cell(row=i, column=int(columns['资源类型'])).value == "数据库" or table.cell(row=1, column=int(
                            columns['资源名称'])).value == "数据库名称":
                        account_detail_database = databaseAccountList.objects.filter(account_name=account_name,
                                                                                     database_name=resource_name)
                        if account_detail_database:
                            account_detail_database = databaseAccountList.objects.get(account_name=account_name,
                                                                                      database_name=resource_name)
                            if account_detail_database.is_in_resource == "否":
                                account_detail_database.is_new = "否"
                                account_detail_database.is_applied = "是"
                        else:
                            account_detail_database = databaseAccountList.objects.filter(account_4AAccount=account_4AAccount,
                                                                                         account_name="",
                                                                                         database_name=resource_name)
                            if account_detail_database:
                                account_detail_database = databaseAccountList.objects.get(account_4AAccount=account_4AAccount,
                                                                                          account_name="",
                                                                                          database_name=resource_name)
                            else:
                                account_detail_database = databaseAccountList()
                                account_detail_database.is_new = "是"
                                account_detail_database.is_applied = "否"
                            account_detail_database.account_name = account_name
                            account_detail_database.database_name = resource_name
                            account_detail_database.database_ip = resource_ip
                            account_detail_database.is_in_resource = "否"
                        if account_detail_database.account_4AAccount in [None, ""] or account_detail_database.account_4AAccount == account_4AAccount:
                            account_detail_database.account_4AAccount = account_4AAccount
                            account_detail_database.is_public = "否"
                        else:
                            account_detail_database.account_4AAccount = account_detail_database.account_4AAccount + "," + account_4AAccount
                            account_detail_database.is_public = "是"
                        if account_type in ["用户具名账号", "普通帐号"] and \
                           staffInfo.objects.filter(Q(staff_name=staff_name) &
                                                    Q(staff_system__icontains=system[0]) &
                                                    Q(staff_status="在职")):
                            account_detail_database.is_to_cancel = "否"
                        elif account_type not in ["用户具名账号", "普通帐号"]:
                            account_detail_database.is_to_cancel = "否"
                            if account_name + "@" not in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                matchList.append(account_name + "@")
                        else:
                            account_detail_database.is_to_cancel = "是"
                        if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                            account_detail_database.staff_name = staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_name
                            if account_type not in ["用户具名账号", "普通帐号"]:
                                if account_name + "@" in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                    matchList.remove(account_name + "@")
                                account_detail_database.is_to_change = "是"
                                if account_detail_database.how_to_change == "":
                                    account_detail_database.how_to_change = "变更账号属性为用户具名账号/普通帐号"
                                elif not re.search("账号属性", account_detail_database.how_to_change):
                                    account_detail_database.how_to_change = account_detail_database.how_to_change + "+变更账号属性为用户具名账号/普通帐号"
                        else:
                            account_detail_database.staff_name = staff_name
                        account_detail_database.account_type = account_type
                        account_detail_database.database_system = system[0]
                        account_detail_database.is_keep = "是"
                        account_detail_database.is_in_4A = "是"
                        if account_detail_database.account_4AAccount in [None, ""] and account_detail_database.is_in_4A == "是":
                            account_detail_database.is_isolated = "是"
                        else:
                            account_detail_database.is_isolated = "否"
                        account_detail_database.save()
                    else:
                        if re.search(system[0], resource_system) or resource_system in system:
                            account_detail_application = applicationAccountList.objects.filter(account_name=account_name)
                            if account_detail_application:
                                account_detail_application = applicationAccountList.objects.get(account_name=account_name)
                                account_detail_application.is_new = "否"
                                account_detail_application.is_applied = "是"
                            else:
                                account_detail_application = applicationAccountList()
                                account_detail_application.account_name = account_name
                                account_detail_application.is_in_resource = "是"
                                account_detail_application.is_new = "是"
                                account_detail_application.is_applied = "否"
                            if account_detail_application.account_4AAccount in [None, ""] or account_detail_application.account_4AAccount == account_4AAccount:
                                account_detail_application.account_4AAccount = account_4AAccount
                                account_detail_application.is_public = "否"
                            else:
                                account_detail_application.account_4AAccount = account_detail_application.account_4AAccount + "," + str(account_4AAccount)
                                account_detail_application.is_public = "是"
                            if account_type in ["用户具名账号", "普通帐号"] and \
                               staffInfo.objects.filter(Q(staff_name=staff_name) &
                                                        Q(staff_system__icontains=system[0]) &
                                                        Q(staff_status="在职")):
                                account_detail_application.is_to_cancel = "否"
                            elif account_type not in ["用户具名账号", "普通帐号"]:
                                account_detail_application.is_to_cancel = "否"
                                if account_name + "@" not in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                    matchList.append(account_name + "@")
                            else:
                                account_detail_application.is_to_cancel = "是"
                            if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                                account_detail_application.staff_name = staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_name
                                if account_type not in ["用户具名账号", "普通帐号"]:
                                    if account_name + "@" in matchList and system[0] in ["PBOSS", "内容计费", "物漫"]:
                                        matchList.remove(account_name + "@")
                                    account_detail_application.is_to_change = "是"
                                    if account_detail_application.how_to_change == "":
                                        account_detail_application.how_to_change = "变更账号属性为用户具名账号/普通帐号"
                                    elif not re.search("账号属性", account_detail_application.how_to_change):
                                        account_detail_application.how_to_change = account_detail_application.how_to_change + "+变更账号属性为用户具名账号/普通帐号"
                            else:
                                account_detail_application.staff_name = staff_name
                            account_detail_application.account_type = account_type
                            account_detail_application.application_system = system[0]
                            account_detail_application.is_keep = "是"
                            account_detail_application.is_in_4A = "是"
                            if account_detail_application.account_4AAccount in [None, ""] and account_detail_application.is_in_4A == "是":
                                account_detail_application.is_isolated = "是"
                            else:
                                account_detail_application.is_isolated = "否"
                            account_detail_application.save()
                else:
                    if table.cell(row=i, column=int(columns['资源类型'])).value == "unix" or table.cell(row=1, column=int(
                            columns['资源名称'])).value == "主机名称":
                        hostAccountList.objects.filter(account_name=account_name, host_system=system[0], host_name=resource_name).delete()
                    elif table.cell(row=i, column=int(columns['资源类型'])).value == "数据库" or table.cell(row=1, column=int(
                            columns['资源名称'])).value == "数据库名称":
                        databaseAccountList.objects.filter(account_name=account_name, database_system=system[0], database_name=resource_name).delete()
                    else:
                        applicationAccountList.objects.filter(account_name=account_name, application_system=system[0]).delete()
                hostAccountList.objects.filter(is_to_change="", host_system=system[0]).update(is_to_change="否")
                databaseAccountList.objects.filter(is_to_change="", database_system=system[0]).update(is_to_change="否")
                applicationAccountList.objects.filter(is_to_change="", application_system=system[0]).update(is_to_change="否")
                # hostAccountList.objects.filter(is_in_resource="否", host_system=system[0]).update(account_name="", account_group="")
                # databaseAccountList.objects.filter(is_in_resource="否", database_system=system[0]).update(account_name="", account_role="")
                # applicationAccountList.objects.filter(is_in_resource="否", application_system=system[0]).update(account_name="")
                print(matchList)
    except:
        import traceback
        traceback.print_exc()
        ret = {'msg': "请上传4A主账号与从账号对应清单！"}
    try:
        application = load_workbook(audit_file_upload + "开户+变更+销户记录.xlsx")
        application_types = application.get_sheet_names()
        for application_type in application_types:
            apply = application.get_sheet_by_name(application_type)
            nrows = apply.max_row
            ncolumns = apply.max_column
            if application_type == "开户":
                for i in range(2, nrows + 1):
                    account_name = apply.cell(row=i, column=4).value
                    resource_name = apply.cell(row=i, column=6).value
                    resource_type = apply.cell(row=i, column=8).value
                    account_privs = apply.cell(row=i, column=9).value
                    if resource_type == "主机":
                        hostAccountList.objects.filter(account_name=account_name, host_name=resource_name,
                                                       account_group=account_privs).update(is_applied="是")
                    elif resource_type == "数据库":
                        account_privs = "+".join(account_privs.split(","))
                        databaseAccountList.objects.filter(account_name=account_name, database_name=resource_name,
                                                           account_role=account_privs).update(is_applied="是")
                    else:
                        applicationAccountList.objects.filter(account_name=account_name,
                                                              account_role=account_privs).update(is_applied="是")
                # hostAccountList.objects.filter(is_new="是", is_applied="否").update(is_to_cancel="是")
                # databaseAccountList.objects.filter(is_new="是", is_applied="否").update(is_to_cancel="是")
                # applicationAccountList.objects.filter(is_new="是", is_applied="否").update(is_to_cancel="是")
            elif application_type == "销户":
                for i in range(2, nrows + 1):
                    account_name = apply.cell(row=i, column=4).value
                    system_name = apply.cell(row=i, column=5).value
                    hostAccountList.objects.filter(account_name=account_name, host_system=system_name).update(is_keep="否")
                    databaseAccountList.objects.filter(account_name=account_name, database_system=system_name).update(is_keep="否")
                    applicationAccountList.objects.filter(account_name=account_name, application_system=system_name).update(is_keep="否")
                hostAccountList.objects.filter(is_keep="否", is_in_4A="否", is_in_resource="否").delete()
                databaseAccountList.objects.filter(is_keep="否", is_in_4A="否", is_in_resource="否").delete()
                applicationAccountList.objects.filter(is_keep="否", is_in_4A="否", is_in_resource="否").delete()
                hostAccountList.objects.filter(is_keep="否", host_system=system[0]).update(is_to_cancel="否")
                databaseAccountList.objects.filter(is_keep="否", database_system=system[0]).update(is_to_cancel="否")
                applicationAccountList.objects.filter(is_keep="否", application_system=system[0]).update(is_to_cancel="否")
    except:
        ret = {'msg': "请上传开户+变更+销户记录！"}
    try:
        csv_file = csv.reader(open(audit_file_upload + "操作日志.csv", 'r'))
        csv_header = next(csv_file)
        columns = {'操作内容': "-1",
                   '操作时间': "-1",
                   '操作人帐号': "-1",
                   '特权/程序账号': "-1",
                   '资源IP': "-1",
                   '是否金库操作': "-1",
                   '资源名称': "-1"}
        for i in range(0, len(csv_header)):
            if csv_header[i] in ['操作内容']:
                columns['操作内容'] = i
            elif csv_header[i] in ['操作日期', '操作时间', '日期']:
                columns['操作时间'] = i
            elif csv_header[i] in ['操作人从帐号', '主帐号名称', '从账号']:
                columns['操作人帐号'] = i
            elif csv_header[i] in ['从帐号名称']:
                columns['特权/程序账号'] = i
            elif csv_header[i] in ['目的设备名称']:
                columns['资源IP'] = i
            elif csv_header[i] in ['是否金库操作']:
                columns['是否金库操作'] = i
            elif csv_header[i] in ['数据库服务名']:
                columns['资源名称'] = i
        result_file = open(audit_file_result + "使用特权+程序账号未备案记录.csv", 'w', newline='')
        result_file_sensitive = open(audit_file_result + "敏感操作未触发金库.csv", 'w', newline='')
        csv_result = csv.writer(result_file)
        csv_result_sensitive = csv.writer(result_file_sensitive)
        csv_result.writerow(csv_header)
        csv_result_sensitive.writerow(csv_header)
        row_num_sum = 0
        row_num_without_application = 0
        row_num_without_authentication = 0
        for row in csv_file:
            row_num_sum += 1
            if re.search("T", row[int(columns['操作时间'])], re.IGNORECASE):
                date = row[int(columns['操作时间'])].split("T")[0]
                time_detail = row[int(columns['操作时间'])].split("T")[1].split(".")[0]
            elif row[int(columns['操作时间'])]:
                date = row[int(columns['操作时间'])].split(" ")[0]
                if not date:
                    continue
                date = datetime.datetime.strptime(date, r'%Y/%m/%d')
                date = datetime.datetime.strftime(date, '%Y-%m-%d')
                if row[int(columns['操作时间'])].split(" ")[1]:
                    time_detail = row[int(columns['操作时间'])].split(" ")[1]
                else:
                    time_detail = row[int(columns['操作时间'])].split(" ")[2]
            else:
                continue
            if len(time_detail) < 7:
                time_detail = time_detail + ":00"
            elif len(time_detail) > 8:
                time_detail = time_detail[0:7]
            operating_time = int(time.mktime(time.strptime(date + " " + time_detail, '%Y-%m-%d %H:%M:%S')))
            personal_account = row[int(columns['操作人帐号'])]
            if int(columns['特权/程序账号']) == -1:
                if row[int(columns['操作内容'])] == "" or not re.search("@", row[int(columns['操作内容'])], re.IGNORECASE):
                    continue
            # print(row[int(columns['操作时间'])] + row[int(columns['操作内容'])])
            if (personal_account + "@") in matchList and int(columns['特权/程序账号']) == -1:
                continue
            else:
                for listrow in matchList:
                    if re.search(listrow, row[int(columns['操作内容'])], re.IGNORECASE) or \
                       row[int(columns['特权/程序账号'])] + "@" == listrow or \
                       row[int(columns['特权/程序账号'])] == listrow:
                        # 匹配进行ssh操作的记录
                        if listrow == "ssh" or re.search("ssh", row[int(columns['操作内容'])], re.IGNORECASE):
                            print(personal_account + ":a:" + row[int(columns['操作内容'])])
                            pattern = re.compile(r'[s{2}][h{1}][ ](.*?)$')
                            label = re.findall(pattern, row[int(columns['操作内容'])])
                            if label and re.search("@", label[0]):
                                privileged_account = label[0].split("@")[0].strip()
                                if re.search(" ", privileged_account):
                                    privileged_account = privileged_account.split(" ")[1]
                                # 匹配ssh XXXX@localhost
                                if label[0].split("@")[1] == "localhost":
                                    label1 = re.findall(re.compile(r'[@](.*?)[:]'), row[int(columns['操作内容'])])
                                    label2 = re.findall(re.compile(r'[@](.*?)[ ]'), row[int(columns['操作内容'])])
                                    if label1:
                                        resource = label1[0].split(" ")[0].split(":")[0].split("[")[0].strip()
                                    elif label2:
                                        resource = label2[0].split(" ")[0].split(":")[0].split("[")[0].strip()
                                    elif int(columns['特权/程序账号']) != -1:
                                        resource = row[int(columns['资源IP'])]
                                    else:
                                        continue
                                # 匹配ssh XXXX@XXXX
                                elif not label[0].split("@")[1].startswith("$"):
                                    resource = label[0].split("@")[1].split(" ")[0].split(":")[0].split("[")[0].strip()
                                else:
                                    continue
                            else:
                                # 匹配ssh XXXX
                                if label and re.findall(re.compile(r'([A-Za-z0-9]*?)[@]'), row[int(columns['操作内容'])]):
                                    if int(columns['特权/程序账号']) == -1:
                                        privileged_account = re.findall(re.compile(r'([A-Za-z0-9]*?)[@]'), row[int(columns['操作内容'])])[0].strip()
                                    else:
                                        privileged_account = row[int(columns['特权/程序账号'])]
                                    resource = label[0].split(" ")[0].split(":")[0].split("[")[0].strip()
                                else:
                                    continue
                        else:
                            pattern1 = re.compile(r'[[](.*?)[ ]')
                            pattern2 = re.compile(r'[@](.*?)[:]')
                            pattern3 = re.compile(r'[@](.*?)[>]')
                            pattern4 = re.compile(r'(.*?)[@]')
                            if listrow == "sudo" or listrow == "root":
                                privileged_account = "root"
                                #print(personal_account + ":b:" + row[int(columns['特权/程序账号'])] + ":" + row[int(columns['操作内容'])])
                            else:
                                privileged_account = listrow.replace("@", "")
                            # 匹配[pboss@PBOSS2-APP3 ~]$
                            label1 = re.findall(pattern1, row[int(columns['操作内容'])])
                            if int(columns['特权/程序账号']) == -1:
                                if label1:
                                    item1 = label1[0].split("@")
                                else:
                                    item1 = []
                                # 匹配pboss@PBOSS3-APP30:/app/pboss3workflow/workflow/workflow-scan/bin>
                                label2 = re.findall(pattern2, row[int(columns['操作内容'])])
                                item2 = re.findall(pattern4, row[int(columns['操作内容'])])
                                # 匹配provision@pbcdb1(2355)>
                                label3 = re.findall(pattern3, row[int(columns['操作内容'])])
                                item3 = re.findall(pattern4, row[int(columns['操作内容'])])
                                if len(item1) == 2 and not item1[1].startswith(".") and not item1[1].startswith("$") and (re.search(privileged_account, item1[0], re.IGNORECASE) or listrow == "sudo"):
                                    if re.search(":", item1[1], re.IGNORECASE):
                                        resource = item1[1].split(":")[0].split("[")[0].strip()
                                    else:
                                        resource = item1[1].split("[")[0].strip()
                                elif label2 and not label2[0].startswith(".") and not label2[0].startswith("$") and (re.search(privileged_account, item2[0], re.IGNORECASE) or listrow == "sudo"):
                                    resource = label2[0].split(" ")[0].split("[")[0].strip()
                                elif label3 and not label3[0].startswith(".") and not label3[0].startswith("$") and (re.search(privileged_account, item3[0], re.IGNORECASE) or listrow == "sudo"):
                                    resource = label3[0].split(" ")[0].split("[")[0].strip()
                                else:
                                    continue
                            else:
                                resource = row[int(columns['资源IP'])]
                        if (privileged_account + "@" not in matchList and system[0] == "PBOSS") or (privileged_account not in matchList and system[0] != "PBOSS"):
                            continue
                        if privileged_account == "I" and system[0] == "PBOSS":
                            privileged_account = "I2000"
                        elif (privileged_account == "P" or privileged_account == "p") and system[0] == "PBOSS":
                            privileged_account = "PBOSS"
                        records = privilegedAccountRecord.objects.filter((Q(resource_name__icontains=resource) &
                                                                          Q(personal_account__icontains=personal_account) &
                                                                          Q(privileged_account__icontains=privileged_account)) |
                                                                         (Q(resource_ip__icontains=resource) &
                                                                          Q(personal_account__icontains=personal_account) &
                                                                          Q(privileged_account__icontains=privileged_account)))
                        if records:
                            for record in records:
                                if operating_time in range(record.start_time, record.end_time):
                                    break
                                elif operating_time not in range(record.start_time, record.end_time) and record == records[len(records) - 1]:
                                    row_num_without_application += 1
                                    if privilegedAccountWithoutRecord.objects.filter(Q(resource__icontains=resource) &
                                                                                     Q(personal_account=personal_account) &
                                                                                     Q(privileged_account=privileged_account) &
                                                                                     Q(operating_time=date)):
                                        pass
                                    else:
                                        without_record = privilegedAccountWithoutRecord()
                                        without_record.resource = resource
                                        without_record.personal_account = personal_account
                                        without_record.privileged_account = privileged_account
                                        without_record.operating_time = date
                                        without_record.save()
                                        csv_result.writerow(row)
                                        break
                        else:
                            row_num_without_application += 1
                            if privilegedAccountWithoutRecord.objects.filter(Q(resource__icontains=resource) &
                                                                             Q(personal_account=personal_account) &
                                                                             Q(privileged_account=privileged_account) &
                                                                             Q(operating_time=date)):
                                break
                            else:
                                without_record = privilegedAccountWithoutRecord()
                                without_record.resource = resource
                                without_record.personal_account = personal_account
                                without_record.privileged_account = privileged_account
                                without_record.operating_time = date
                                without_record.save()
                                csv_result.writerow(row)
                                break
                                # print(resource + " " + personal_account + " " + privileged_account + " " + str(operating_time))
            sensitive_items = sensitiveInformation.objects.filter(system=system[0])
            for sensitive_item in sensitive_items:
                pattern1 = re.compile(r'[_a-zA-Z]+' + sensitive_item.sensitive_item, re.I)
                pattern2 = re.compile(sensitive_item.sensitive_item + r'[_a-zA-Z]+', re.I)
                if system[0] in ["PBOSS", "物漫"]:
                    if re.search(sensitive_item.sensitive_item, row[int(columns['操作内容'])], re.IGNORECASE) and row[int(columns['是否金库操作'])] == "否":
                        csv_result_sensitive.writerow(row)
                        break
                    else:
                        continue
                else:
                    print(re.findall(pattern1, row[int(columns['操作内容'])]))
                    print(re.findall(pattern2, row[int(columns['操作内容'])]))
                    if re.search(sensitive_item.sensitive_item, row[int(columns['操作内容'])], re.IGNORECASE) and \
                       not re.findall(pattern1, row[int(columns['操作内容'])]) and \
                       not re.findall(pattern2, row[int(columns['操作内容'])]):
                        date = row[int(columns['操作时间'])].split(" ")[0]
                        if not date or len(date) > 10:
                            continue
                        date = datetime.datetime.strptime(date, r'%Y/%m/%d')
                        date = datetime.datetime.strftime(date, '%Y-%m-%d')
                        if row[int(columns['操作时间'])].split(" ")[1]:
                            time_detail = row[int(columns['操作时间'])].split(" ")[1]
                        else:
                            time_detail = row[int(columns['操作时间'])].split(" ")[2]
                        if len(time_detail) < 7:
                            time_detail = time_detail + ":00"
                        operating_time = int(time.mktime(time.strptime(date + " " + time_detail, '%Y-%m-%d %H:%M:%S')))
                        operating_account = row[int(columns['操作人帐号'])]
                        resource_name = row[int(columns['资源名称'])]
                        records = treasuryAuthenticated.objects.filter((Q(operating_account=operating_account) &
                                                                        Q(resource_name=resource_name) &
                                                                        Q(sensitive_resource__icontains=sensitive_item.sensitive_item)))
                        if records:
                            for record in records:
                                if operating_time not in range(record.start_time, record.end_time) and record != records[len(records) - 1]:
                                    # print("a:" + row[int(columns['操作内容'])])
                                    continue
                                elif operating_time in range(record.start_time, record.end_time):
                                    # print("b:" + row[int(columns['操作内容'])])
                                    break
                                elif operating_time not in range(record.start_time, record.end_time) and record == records[len(records) - 1]:
                                    print("c:" + row[int(columns['操作内容'])] + str(operating_time) + operating_account + ":" + sensitive_item.sensitive_item)
                                    row_num_without_authentication += 1
                                    csv_result_sensitive.writerow(row)
                                    break
                                break
                        else:
                            print("d:" + row[int(columns['操作内容'])] + str(operating_time) + operating_account + ":" + sensitive_item.sensitive_item)
                            row_num_without_authentication += 1
                            csv_result_sensitive.writerow(row)
                            break
        result_file.close()
        result_file_sensitive.close()
        log_num = open(audit_file_result + "log_num.txt", "w")
        log_num.write(str(row_num_sum) + "," + str(row_num_without_application) + "," + str(row_num_without_authentication))
        log_num.close()
    except:
        import traceback
        traceback.print_exc()
        ret = {'msg': "请上传操作日志！"}
    account_audit_result = load_workbook(audit_file_result + "账号审计明细模板.xlsx")
    account_audit_types = account_audit_result.get_sheet_names()
    for account_audit_type in account_audit_types:
        account_audit_detail = account_audit_result.get_sheet_by_name(account_audit_type)
        row = 2
        if account_audit_type == "孤立账号":
            host_isolated_accounts = hostAccountList.objects.filter(host_system=system[0], is_isolated="是")
            database_isolated_accounts = databaseAccountList.objects.filter(database_system=system[0], is_isolated="是")
            application_isolated_accounts = applicationAccountList.objects.filter(application_system=system[0], is_isolated="是")
            for host_isolated_account in host_isolated_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_isolated_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_isolated_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_isolated_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_isolated_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_isolated_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_isolated_account.account_type)
                account_audit_detail.cell(row=row, column=8, value="否")
                account_audit_detail.cell(row=row, column=9, value=host_isolated_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for database_isolated_account in database_isolated_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_isolated_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_isolated_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_isolated_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_isolated_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_isolated_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_isolated_account.account_type)
                account_audit_detail.cell(row=row, column=8, value="否")
                account_audit_detail.cell(row=row, column=9, value=database_isolated_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for application_isolated_account in application_isolated_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_isolated_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_isolated_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_isolated_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_isolated_account.account_type)
                account_audit_detail.cell(row=row, column=8, value="否")
                account_audit_detail.cell(row=row, column=9, value=application_isolated_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
        elif account_audit_type == "公共账号":
            host_public_accounts = hostAccountList.objects.filter(host_system=system[0], is_public="是")
            database_public_accounts = databaseAccountList.objects.filter(database_system=system[0], is_public="是")
            application_public_accounts = applicationAccountList.objects.filter(application_system=system[0], is_public="是")
            for host_public_account in host_public_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_public_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_public_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_public_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_public_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_public_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_public_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_public_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value="是")
                account_audit_detail.cell(row=row, column=10, value=host_public_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=11, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for database_public_account in database_public_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_public_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_public_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_public_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_public_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_public_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_public_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_public_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value="是")
                account_audit_detail.cell(row=row, column=10, value=database_public_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=11, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for application_public_account in application_public_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_public_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_public_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_public_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_public_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_public_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value="是")
                account_audit_detail.cell(row=row, column=10, value=application_public_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=11, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
        elif account_audit_type == "4A侧存在资源侧不存在账号":
            host_inconsistent1_accounts = hostAccountList.objects.filter(host_system=system[0], is_in_4A="是", is_in_resource="否")
            database_inconsistent1_accounts = databaseAccountList.objects.filter(database_system=system[0], is_in_4A="是", is_in_resource="否")
            application_inconsistent1_accounts = applicationAccountList.objects.filter(application_system=system[0], is_in_4A="是", is_in_resource="否")
            for host_inconsistent1_account in host_inconsistent1_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_inconsistent1_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_inconsistent1_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_inconsistent1_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_inconsistent1_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_inconsistent1_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_inconsistent1_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_inconsistent1_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_inconsistent1_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                account_audit_detail.cell(row=row, column=12, value="否")
                row += 1
            for database_inconsistent1_account in database_inconsistent1_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_inconsistent1_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_inconsistent1_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_inconsistent1_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_inconsistent1_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_inconsistent1_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_inconsistent1_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_inconsistent1_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_inconsistent1_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                account_audit_detail.cell(row=row, column=12, value="否")
                row += 1
            for application_inconsistent1_account in application_inconsistent1_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_inconsistent1_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_inconsistent1_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_inconsistent1_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_inconsistent1_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_inconsistent1_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=application_inconsistent1_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                account_audit_detail.cell(row=row, column=12, value="否")
                row += 1
        elif account_audit_type == "资源侧存在4A侧不存在账号":
            host_inconsistent2_accounts = hostAccountList.objects.filter(host_system=system[0], is_in_4A="否", is_in_resource="是")
            database_inconsistent2_accounts = databaseAccountList.objects.filter(database_system=system[0], is_in_4A="否", is_in_resource="是")
            application_inconsistent2_accounts = applicationAccountList.objects.filter(application_system=system[0], is_in_4A="否", is_in_resource="是")
            for host_inconsistent2_account in host_inconsistent2_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_inconsistent2_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_inconsistent2_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_inconsistent2_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_inconsistent2_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_inconsistent2_account.account_group)
                account_audit_detail.cell(row=row, column=7, value="否")
                account_audit_detail.cell(row=row, column=8, value="是")
                account_audit_detail.cell(row=row, column=9, value=host_inconsistent2_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for database_inconsistent2_account in database_inconsistent2_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_inconsistent2_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_inconsistent2_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_inconsistent2_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_inconsistent2_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_inconsistent2_account.account_role)
                account_audit_detail.cell(row=row, column=7, value="否")
                account_audit_detail.cell(row=row, column=8, value="是")
                account_audit_detail.cell(row=row, column=9, value=database_inconsistent2_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
            for application_inconsistent2_account in application_inconsistent2_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_inconsistent2_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_inconsistent2_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_inconsistent2_account.account_role)
                account_audit_detail.cell(row=row, column=7, value="否")
                account_audit_detail.cell(row=row, column=8, value="是")
                account_audit_detail.cell(row=row, column=9, value=application_inconsistent2_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                row += 1
        elif account_audit_type == "需变更主机账号主组":
            host_change_group_accounts = hostAccountList.objects.filter(Q(host_system=system[0]) & Q(how_to_change__icontains="变更主组"))
            for host_change_group_account in host_change_group_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_change_group_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_change_group_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_change_group_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_change_group_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_change_group_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_change_group_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_change_group_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_change_group_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value=host_change_group_account.how_to_change)
                row += 1
        elif account_audit_type == "需变更数据库账号角色":
            database_change_role_accounts = databaseAccountList.objects.filter(Q(database_system=system[0]) & Q(how_to_change__icontains="变更角色权限"))
            for database_change_role_account in database_change_role_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_change_role_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_change_role_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_change_role_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_change_role_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_change_role_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_change_role_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_change_role_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_change_role_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value=database_change_role_account.how_to_change)
                row += 1
        elif account_audit_type == "需变更账号属性":
            host_change_type_accounts = hostAccountList.objects.filter(Q(host_system=system[0]) & Q(how_to_change__icontains="变更账号属性"))
            database_change_type_accounts = databaseAccountList.objects.filter(Q(database_system=system[0]) & Q(how_to_change__icontains="变更账号属性"))
            application_change_type_accounts = applicationAccountList.objects.filter(Q(application_system=system[0]) & Q(how_to_change__icontains="变更账号属性"))
            for host_change_type_account in host_change_type_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_change_type_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_change_type_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_change_type_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_change_type_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_change_type_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_change_type_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_change_type_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_change_type_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value=host_change_type_account.how_to_change)
                row += 1
            for database_change_type_account in database_change_type_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_change_type_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_change_type_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_change_type_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_change_type_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_change_type_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_change_type_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_change_type_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_change_type_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value=database_change_type_account.how_to_change)
                row += 1
            for application_change_type_account in application_change_type_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_change_type_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_change_type_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_change_type_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_change_type_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_change_type_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=application_change_type_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value=application_change_type_account.how_to_change)
                row += 1
        elif account_audit_type == "已销户漏删账号":
            host_miss_delete_accounts = hostAccountList.objects.filter(host_system=system[0], is_keep="否")
            database_miss_delete_accounts = databaseAccountList.objects.filter(database_system=system[0], is_keep="否")
            application_miss_delete_accounts = applicationAccountList.objects.filter(application_system=system[0], is_keep="否")
            for host_miss_delete_account in host_miss_delete_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_miss_delete_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_miss_delete_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_miss_delete_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_miss_delete_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_miss_delete_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_miss_delete_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_miss_delete_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_miss_delete_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
            for database_miss_delete_account in database_miss_delete_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_miss_delete_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_miss_delete_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_miss_delete_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_miss_delete_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_miss_delete_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_miss_delete_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_miss_delete_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_miss_delete_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
            for application_miss_delete_account in application_miss_delete_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_miss_delete_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_miss_delete_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_miss_delete_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_miss_delete_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_miss_delete_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=application_miss_delete_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
        elif account_audit_type == "需提单销户":
            host_to_cancel_accounts = hostAccountList.objects.filter(host_system=system[0], is_to_cancel="是")
            database_to_cancel_accounts = databaseAccountList.objects.filter(database_system=system[0], is_to_cancel="是")
            application_to_cancel_accounts = applicationAccountList.objects.filter(application_system=system[0], is_to_cancel="是")
            for host_to_cancel_account in host_to_cancel_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_to_cancel_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_to_cancel_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_to_cancel_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_to_cancel_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_to_cancel_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_to_cancel_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_to_cancel_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_to_cancel_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                row += 1
            for database_to_cancel_account in database_to_cancel_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_to_cancel_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_to_cancel_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_to_cancel_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_to_cancel_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_to_cancel_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_to_cancel_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_to_cancel_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_to_cancel_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                row += 1
            for application_to_cancel_account in application_to_cancel_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_to_cancel_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_to_cancel_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_to_cancel_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_to_cancel_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_to_cancel_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=application_to_cancel_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="是")
                row += 1
        elif account_audit_type == "新增账号无开户记录":
            host_new_without_application_accounts = hostAccountList.objects.filter(host_system=system[0], is_keep="是", is_applied="否", is_to_cancel="否")
            database_new_without_application_accounts = databaseAccountList.objects.filter(database_system=system[0], is_keep="是", is_applied="否", is_to_cancel="否")
            application_new_without_application_accounts = applicationAccountList.objects.filter(application_system=system[0], is_keep="是", is_applied="否", is_to_cancel="否")
            for host_new_without_application_account in host_new_without_application_accounts:
                account_audit_detail.cell(row=row, column=1, value="主机")
                account_audit_detail.cell(row=row, column=2, value=host_new_without_application_account.host_ip)
                account_audit_detail.cell(row=row, column=3, value=host_new_without_application_account.host_name)
                account_audit_detail.cell(row=row, column=4, value=host_new_without_application_account.host_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=host_new_without_application_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=host_new_without_application_account.account_group)
                account_audit_detail.cell(row=row, column=7, value=host_new_without_application_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=host_new_without_application_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=host_new_without_application_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
            for database_new_without_application_account in database_new_without_application_accounts:
                account_audit_detail.cell(row=row, column=1, value="数据库")
                account_audit_detail.cell(row=row, column=2, value=database_new_without_application_account.database_ip)
                account_audit_detail.cell(row=row, column=3, value=database_new_without_application_account.database_name)
                account_audit_detail.cell(row=row, column=4, value=database_new_without_application_account.database_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=database_new_without_application_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=database_new_without_application_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=database_new_without_application_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=database_new_without_application_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=database_new_without_application_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
            for application_new_without_application_account in application_new_without_application_accounts:
                account_audit_detail.cell(row=row, column=1, value="应用")
                account_audit_detail.cell(row=row, column=4, value=application_new_without_application_account.application_system)
                account_name = account_audit_detail.cell(row=row, column=5, value=application_new_without_application_account.account_name)
                account_audit_detail.cell(row=row, column=6, value=application_new_without_application_account.account_role)
                account_audit_detail.cell(row=row, column=7, value=application_new_without_application_account.account_type)
                account_audit_detail.cell(row=row, column=8, value=application_new_without_application_account.account_4AAccount)
                account_audit_detail.cell(row=row, column=9, value=application_new_without_application_account.staff_name)
                account_name = account_name.value
                if staffInfo.objects.filter(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])):
                    account_audit_detail.cell(row=row, column=10, value=staffInfo.objects.get(Q(staff_account=account_name) & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains=account_name + ",") & Q(staff_system__icontains=system[0]) |
                                            Q(staff_account__icontains="," + account_name) & Q(staff_system__icontains=system[0])).staff_group)
                account_audit_detail.cell(row=row, column=11, value="否")
                row += 1
    account_audit_result.save(audit_file_result + "账号审计明细.xlsx")
    try:
        csv_file = open(audit_file_result + "使用特权+程序账号未备案记录.csv", encoding='ansi')
        writer = pd.ExcelWriter(audit_file_result + "使用特权+程序账号未备案.xlsx")
        pd.read_csv(csv_file).to_excel(writer, sheet_name="使用特权+程序账号未备案", index=False, encoding='ansi')
    except:
        import traceback
        traceback.print_exc()
    writer.save()
    writer.close()
    try:
        csv_file_sensitive = open(audit_file_result + "敏感操作未触发金库.csv", encoding='ansi')
        writer = pd.ExcelWriter(audit_file_result + "敏感操作未触发金库.xlsx")
        pd.read_csv(csv_file_sensitive).to_excel(writer, sheet_name="敏感操作未触发金库", index=False, encoding='ansi')
    except:
        import traceback
        traceback.print_exc()
    writer.save()
    writer.close()
    v = json.dumps(ret)
    return HttpResponse(v)


def exportAccountAuditResult(request):
    wb = load_workbook(audit_file_result + "账号审计明细.xlsx")
    wb.create_sheet(title="使用特权+程序账号未备案", index=0)
    wb.create_sheet(title="敏感操作未触发金库", index=1)
    temp = load_workbook(audit_file_result + "使用特权+程序账号未备案.xlsx")
    temp_sensitive = load_workbook(audit_file_result + "敏感操作未触发金库.xlsx")
    sheet_old = temp.get_sheet_by_name("使用特权+程序账号未备案")
    sheet_old_sensitive = temp_sensitive.get_sheet_by_name("敏感操作未触发金库")
    sheet_new = wb.get_sheet_by_name("使用特权+程序账号未备案")
    sheet_new_sensitive = wb.get_sheet_by_name("敏感操作未触发金库")
    max_row = sheet_old.max_row
    max_row_sensitive = sheet_old_sensitive.max_row
    max_column = sheet_old.max_column
    max_column_sensitive = sheet_old_sensitive.max_column
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            value = sheet_old.cell(row=i, column=j).value
            sheet_new.cell(row=i, column=j, value=value)
    for i in range(1, max_row_sensitive + 1):
        for j in range(1, max_column_sensitive + 1):
            value = sheet_old_sensitive.cell(row=i, column=j).value
            sheet_new_sensitive.cell(row=i, column=j, value=value)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    time = datetime.datetime.now().strftime('%Y-%m-%d')
    file_name = '账号审计明细%s.xlsx' % time
    file_name = urlquote(file_name)
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response
